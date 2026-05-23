#!/usr/bin/env python3
"""
NSAT — Firestore seed script.

Seeds the `questions` and `tests` collections from the filled Excel
question bank, for the nsat-niu-app Firebase project.

Supports both MCQ and Short Answer question types.

WHAT IT DOES
  - Reads the Questions sheet  -> writes `questions` documents
  - Reads the Test Config sheet -> writes `tests` documents
  - MCQ: maps letter answer (A/B/C/D) to correctAnswerIndex (0-3)
  - Short Answer: sets type='shortAnswer', minWords, maxWords, no scoring fields
  - Maps the Excel's display course name to a canonical course KEY

BEFORE RUNNING
  1. pip install openpyxl firebase-admin
  2. Download a service account key for nsat-niu-app:
     Firebase console -> Project settings -> Service accounts ->
     Generate new private key. Save as serviceAccountKey.json next to
     this script. NEVER commit that file.

RUN
  python seed_firestore.py NSAT_QuestionBank_Template.xlsx
  Add --dry-run to print what would be written without touching Firestore.
  Add --clear to delete existing questions and tests before seeding.
"""

import sys
import argparse
import openpyxl

# --- Course → School Paper mapping ----------------------------------------
# NPF returns specific courses; question papers are per SCHOOL (UG/PG split).
# Source: https://niu.edu.in/courses-fee-structure-for-2026-27/
#
# Paper keys: soahs_ug, soahs_pg, son, set_ug, set_pg, sbm_ug, sbm_pg,
#             solla_ug, solla_pg, sjmc, sos_ug, sos_pg, sola, sofad, soe, sop
COURSE_KEY_MAP = {
    # SOAHS UG: School of Allied Health & Care Sciences (Undergraduate)
    "BPT": "soahs_ug", "B.Optom": "soahs_ug", "B. Optom": "soahs_ug",
    "BMLS": "soahs_ug", "BMLT": "soahs_ug", "BMRIT": "soahs_ug",
    "B.Sc-RIT": "soahs_ug", "BRIT": "soahs_ug", "B. AOTT": "soahs_ug",
    "B.Sc-OTT": "soahs_ug", "B.Sc (CCT)": "soahs_ug", "B.Sc CCT": "soahs_ug",
    "B.Sc- CCT": "soahs_ug", "BND": "soahs_ug", "B.Sc-N&D": "soahs_ug",
    "B.Sc (Cardiac Care Technology)": "soahs_ug", "BOT": "soahs_ug",
    "B.Sc- Opt": "soahs_ug", "SOAHS UG": "soahs_ug", "soahs_ug": "soahs_ug",

    # SOAHS PG: School of Allied Health & Care Sciences (Postgraduate)
    "MPT": "soahs_pg", "MMLS": "soahs_pg", "MMLT": "soahs_pg",
    "MMRIT": "soahs_pg", "MPH": "soahs_pg", "PGDEMS": "soahs_pg",
    "soahs_pg": "soahs_pg",

    # SON: School of Nursing
    "GNM": "son", "B.Sc Nursing": "son", "B.Sc-N": "son",
    "B.Sc. Nursing": "son", "B.Sc (Nursing)": "son", "ANM": "son",
    "son": "son",

    # SET UG: School of Engineering & Technology (Undergraduate)
    "B.Tech": "set_ug", "B.Tech / Engineering": "set_ug",
    "B.Tech (CSE)": "set_ug", "B.Tech (AI & ML)": "set_ug",
    "B.Tech (Data Science)": "set_ug", "B.Tech (Cyber Security)": "set_ug",
    "B.Tech (Robotics)": "set_ug", "B.Tech (Biotechnology)": "set_ug",
    "B.Tech (ME)": "set_ug", "B.Tech (CE)": "set_ug", "B.Tech (EE)": "set_ug",
    "B.Tech (ECE)": "set_ug", "B.Tech (IT)": "set_ug",
    "B.Tech (Semiconductor)": "set_ug", "B.Tech (Mechatronics)": "set_ug",
    "B.Tech Lateral Entry": "set_ug",
    "Diploma (Electrical)": "set_ug", "Diploma (Mechanical)": "set_ug",
    "Diploma (Civil)": "set_ug", "Diploma (CSE)": "set_ug",
    "btech": "set_ug", "set_ug": "set_ug",

    # SET PG: School of Engineering & Technology (Postgraduate)
    "M.Tech": "set_pg", "M.Tech (CSE)": "set_pg", "M.Tech (ME)": "set_pg",
    "M.Tech (CE)": "set_pg", "M.Tech (Biotechnology)": "set_pg",
    "M.Tech (EE)": "set_pg", "set_pg": "set_pg",

    # SBM UG: School of Business Management (Undergraduate)
    "BBA": "sbm_ug", "BBA-HHM": "sbm_ug", "BBA-AV": "sbm_ug",
    "BBA (Hons)": "sbm_ug", "B.Com": "sbm_ug", "B.Com (Hons.)": "sbm_ug",
    "B. Com (Hons.)": "sbm_ug", "BFSI": "sbm_ug",
    "Bachelor of Business Administration": "sbm_ug",
    "Bachelor of Commerce": "sbm_ug", "sbm_ug": "sbm_ug",

    # SBM PG: School of Business Management (Postgraduate)
    "MBA": "sbm_pg", "MBA (Finance)": "sbm_pg", "MBA (Marketing)": "sbm_pg",
    "MBA (HR)": "sbm_pg", "MBA-HHM": "sbm_pg", "MBA-PM": "sbm_pg",
    "MBA-ELITE": "sbm_pg", "PGDM": "sbm_pg", "M.Com": "sbm_pg",
    "M. Com": "sbm_pg", "Master of Business Administration": "sbm_pg",
    "sbm_pg": "sbm_pg",

    # SOLLA UG: School of Law & Legal Affairs (5yr integrated)
    "BA LLB": "solla_ug", "BA LLB (Hons.)": "solla_ug",
    "BBA LLB": "solla_ug", "BBA LLB.": "solla_ug",
    "BBA LLB (Hons.)": "solla_ug", "solla_ug": "solla_ug",

    # SOLLA PG: School of Law & Legal Affairs (LLB 3yr / LLM)
    "LLB": "solla_pg", "LLM": "solla_pg", "solla_pg": "solla_pg",

    # SJMC: School of Journalism & Mass Communication
    "BA-JMC": "sjmc", "BA (Journalism & Mass Communication)": "sjmc",
    "MA-JMC": "sjmc", "MA (Journalism & Mass Communication)": "sjmc",
    "sjmc": "sjmc",

    # SOS UG: School of Sciences (Undergraduate)
    "B.Sc": "sos_ug", "B.Sc (Biotechnology)": "sos_ug",
    "B.Sc (Microbiology)": "sos_ug", "B.Sc (Agriculture)": "sos_ug",
    "B.Sc-AG": "sos_ug", "B.Sc (Forensic Science)": "sos_ug",
    "B.Sc (IT)": "sos_ug", "B.Sc (CS)": "sos_ug",
    "BCA": "sos_ug", "BCA (Hons.)": "sos_ug",
    "BCA (AI/ML)": "sos_ug", "BCA (Data Science)": "sos_ug",
    "BCA (Cyber Security)": "sos_ug", "sos_ug": "sos_ug",

    # SOS PG: School of Sciences (Postgraduate)
    "M.Sc": "sos_pg", "M.Sc (Biotechnology)": "sos_pg",
    "M.Sc (Microbiology)": "sos_pg", "M.Sc (Agriculture)": "sos_pg",
    "M.Sc (Forensic Science)": "sos_pg", "M.Sc (IT)": "sos_pg",
    "M.Sc (CS)": "sos_pg", "MCA": "sos_pg", "sos_pg": "sos_pg",

    # SOLA: School of Liberal Arts
    "BA (English)": "sola", "BA (Hons) English": "sola",
    "BA (Psychology)": "sola", "BA (Hons) Psychology": "sola",
    "BA (Sociology)": "sola", "BA (Hons) Sociology": "sola",
    "BA (Political Science)": "sola", "BA (Geography)": "sola",
    "BA (International Relations)": "sola",
    "MA (English)": "sola", "MA (Psychology)": "sola",
    "MA (Geography)": "sola", "sola": "sola",

    # SOFAD: School of Fine Arts & Design (+ Architecture)
    "BFA": "sofad", "BFA (Animation & VFX)": "sofad",
    "BID": "sofad", "B.Des": "sofad", "B.Interior": "sofad",
    "MFA": "sofad", "sofad": "sofad",

    # SOE: School of Education
    "B.Ed": "soe", "B.Ed.": "soe", "MA (Education)": "soe",
    "MA Education": "soe", "soe": "soe",

    # SOP: School of Pharmacy
    "B.Pharm": "sop", "B.Pharma": "sop", "D.Pharm": "sop",
    "D.Pharma": "sop", "M.Pharm": "sop", "sop": "sop",
}

LETTER_TO_INDEX = {"A": 0, "B": 1, "C": 2, "D": 3}

# Header row is row 1, instructions row is row 2, data starts row 3.
DATA_START_ROW = 3


def course_key(display_name):
    name = (display_name or "").strip()
    if name in COURSE_KEY_MAP:
        return COURSE_KEY_MAP[name]
    # Try case-insensitive
    for k, v in COURSE_KEY_MAP.items():
        if k.lower() == name.lower():
            return v
    # Fallback: lowercase slug
    return name.lower().replace(" ", "_").replace("/", "_")


def parse_questions(workbook):
    """Returns a list of dicts ready for the `questions` collection."""
    ws = workbook["Questions"]
    out = []
    course_counters = {}

    for i, row in enumerate(
        ws.iter_rows(min_row=DATA_START_ROW, values_only=True)
    ):
        # Columns: type, course, topic, question_text, opt_a-d, correct, min_words, max_words
        if len(row) < 9:
            continue

        q_type = str(row[0] or "").strip().lower()
        course_raw = row[1]
        topic = row[2]
        qtext = row[3]
        opt_a, opt_b, opt_c, opt_d = row[4], row[5], row[6], row[7]
        correct = row[8]
        min_words = row[9] if len(row) > 9 else None
        max_words = row[10] if len(row) > 10 else None

        excel_row = DATA_START_ROW + i

        if course_raw in (None, ""):
            continue  # blank row
        if qtext in (None, ""):
            continue  # blank question

        ckey = course_key(course_raw)
        topic_str = str(topic).strip() if topic not in (None, "") else ""

        if q_type == "shortanswer":
            # Short answer question — ungraded
            doc = {
                "text": str(qtext).strip(),
                "type": "shortAnswer",
                "options": [],
                "correctAnswerIndex": -1,
                "correctAnswerTexts": [],
                "minWords": int(min_words) if min_words else 100,
                "maxWords": int(max_words) if max_words else 150,
                "course": ckey,
                "topic": topic_str,
            }
            
            course_counters[ckey] = course_counters.get(ckey, 0) + 1
            doc["_doc_id"] = f"{ckey}q{course_counters[ckey]:03d}"
            
            out.append(doc)

        else:
            # MCQ (default) — scored
            letter = (str(correct) if correct else "").strip().upper()
            if letter not in LETTER_TO_INDEX:
                raise ValueError(
                    f"Questions row {excel_row}: correct_option is '{correct}', "
                    f"expected one of A/B/C/D."
                )

            options = [opt_a, opt_b, opt_c, opt_d]
            if any(o in (None, "") for o in options):
                raise ValueError(
                    f"Questions row {excel_row}: one or more options are blank."
                )

            doc = {
                "text": str(qtext).strip(),
                "type": "multipleChoice",
                "options": [str(o).strip() for o in options],
                "correctAnswerIndex": LETTER_TO_INDEX[letter],
                "correctAnswerTexts": [],
                "minWords": 0,
                "maxWords": 0,
                "course": ckey,
                "topic": topic_str,
            }
            
            course_counters[ckey] = course_counters.get(ckey, 0) + 1
            doc["_doc_id"] = f"{ckey}q{course_counters[ckey]:03d}"
            
            out.append(doc)

    return out


def parse_tests(workbook):
    """Returns a list of dicts ready for the `tests` collection."""
    ws = workbook["Test Config"]
    out = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        (course_raw, title, qcount, duration, marks,
         neg_marking, neg_marks, published) = row[:8]
        # Optional 9th column: show_results (defaults to Yes)
        show_results_raw = row[8] if len(row) > 8 and row[8] is not None else "yes"
        if course_raw in (None, ""):
            continue

        try:
            int(qcount)
        except ValueError:
            continue  # Skip header rows

        excel_row = 2 + i

        def yes_no(v, field):
            s = str(v).strip().lower()
            if s in ("yes", "true", "1"):
                return True
            if s in ("no", "false", "0"):
                return False
            raise ValueError(
                f"Test Config row {excel_row}: {field} is '{v}', "
                f"expected yes/no."
            )

        out.append({
            "title": str(title).strip(),
            "course": course_key(course_raw),
            "questionCount": int(qcount),
            "durationMinutes": int(duration),
            "marksPerQuestion": float(marks),
            "negativeMarking": yes_no(neg_marking, "negative_marking"),
            "negativeMarksPerWrong": float(neg_marks),
            "isPublished": yes_no(published, "published"),
            "showResults": yes_no(show_results_raw, "show_results"),
        })
    return out


def validate(questions, tests):
    """Cross-checks the two sheets before anything is written."""
    q_courses = {q["course"] for q in questions}
    t_courses = {t["course"] for t in tests}

    missing_test = q_courses - t_courses
    if missing_test:
        raise ValueError(
            f"Questions exist for course(s) {missing_test} but no test "
            f"is configured for them."
        )

    for t in tests:
        available = sum(1 for q in questions if q["course"] == t["course"])
        if available < t["questionCount"]:
            print(
                f"WARNING: Test '{t['title']}' wants {t['questionCount']} questions "
                f"but only {available} exist for course '{t['course']}'. Proceeding anyway."
            )


def clear_collections(db):
    """Delete all documents in questions and tests collections."""
    for coll_name in ("questions", "tests"):
        docs = db.collection(coll_name).list_documents()
        batch = db.batch()
        count = 0
        for doc in docs:
            batch.delete(doc)
            count += 1
            if count % 450 == 0:
                batch.commit()
                batch = db.batch()
        if count % 450 != 0:
            batch.commit()
        print(f"  Deleted {count} document(s) from '{coll_name}'")


def main():
    parser = argparse.ArgumentParser(description="Seed NSAT Firestore.")
    parser.add_argument("excel", help="Path to the filled question-bank xlsx")
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Print what would be written without touching Firestore",
    )
    parser.add_argument(
        "--clear", action="store_true",
        help="Delete existing questions and tests before seeding",
    )
    parser.add_argument(
        "--key", default="serviceAccountKey.json",
        help="Path to the Firebase service account key",
    )
    args = parser.parse_args()

    workbook = openpyxl.load_workbook(args.excel, data_only=True)
    questions = parse_questions(workbook)
    tests = parse_tests(workbook)
    validate(questions, tests)

    mcq_count = sum(1 for q in questions if q["type"] == "multipleChoice")
    sa_count = sum(1 for q in questions if q["type"] == "shortAnswer")

    print(f"Parsed {len(questions)} question(s): {mcq_count} MCQ + {sa_count} short answer")
    print(f"Parsed {len(tests)} test(s):")
    for t in tests:
        n = sum(1 for q in questions if q["course"] == t["course"])
        print(f"  - '{t['title']}' (course={t['course']}, "
              f"published={t['isPublished']}) <- {n} questions")

    if args.dry_run:
        print("\n--dry-run: nothing written.")
        print("\nSample MCQ document:")
        mcq = next((q for q in questions if q["type"] == "multipleChoice"), None)
        if mcq:
            print(mcq)
        sa = next((q for q in questions if q["type"] == "shortAnswer"), None)
        if sa:
            print("\nSample short answer document:")
            print(sa)
        print("\nSample test document:")
        print(tests[0])
        return

    # --- Real write path -------------------------------------------------
    import firebase_admin
    from firebase_admin import credentials, firestore

    cred = credentials.Certificate(args.key)
    firebase_admin.initialize_app(cred)
    db = firestore.client()

    if args.clear:
        print("\nClearing existing data...")
        clear_collections(db)

    # Batched writes (Firestore batch limit is 500 ops).
    batch = db.batch()
    ops = 0
    for q in questions:
        doc_id = q.pop("_doc_id")
        ref = db.collection("questions").document(doc_id)
        batch.set(ref, q)
        ops += 1
        if ops == 450:
            batch.commit()
            batch = db.batch()
            ops = 0
    for t in tests:
        ref = db.collection("tests").document()  # auto-id
        batch.set(ref, t)
        ops += 1
    if ops:
        batch.commit()

    print(f"\nDone. Wrote {len(questions)} questions ({mcq_count} MCQ + "
          f"{sa_count} short answer) and {len(tests)} tests to Firestore.")


if __name__ == "__main__":
    main()